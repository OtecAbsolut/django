from django.test import TestCase
import json
import os
import openpyxl
# Create your tests here.

# # подгружаем данные из json объекта
# file = open('db_json.json', encoding='utf-8')
# db_json = json.load(file)
# file.close()
#
# # Ищем в словаре данные для menu, product_menu и.т.д.
# for element in db_json:
#     for key, value in element.items():
#         if key == 'menu':
#             menu = value
#         if  key == 'product_menu':
#             product_menu = value
#
# content = {
#         'menu': menu,
#         'product_menu': product_menu,
# }

# Подгрузка базы из exel
FILE_NAME = '/home/novozhilovsv/PROJECT/Учебные материалы/Django/Less4/Данные /готовые фото/exel_load.xlsx'

#  функция загркузки продуктов xlsx файла
def load_products_xlsx(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb['products']
    return sheet

# функция подсчета строк и столбцов в табличке
def cell_coint(sheet):
    i = 1  # определяем число столцов в табице
    j = 1  # определяем число строк
    while True:
        field_name = sheet.cell(row=1, column=i).value
        if not field_name is None:
            i += 1
        else:
            while True:
                row_value = sheet.cell(row=j, column=1).value
                if not row_value is None:
                    j += 1
                else:
                    j = j-1
                    break
            i = i-1
            return i, j

def create_db(sheet, i, j):
    products_all = []
    product = []
    new_db = []
    for n1 in range(1, j+1):
        for n2 in range(1, i+1):
            value = sheet.cell(row=n1, column=n2).value
            product.append(value)
        products_all.append(product)
        product = []
    for index, element in enumerate(products_all):
        if index>0:
            dictionary = dict(zip(products_all[0], element))
            new_db.append(dictionary)
    return new_db

xlsx = load_products_xlsx(FILE_NAME)
i, j = cell_coint(xlsx)
new_db = create_db(xlsx, i, j)


print(new_db)



