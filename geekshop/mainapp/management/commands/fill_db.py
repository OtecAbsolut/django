from django.core.management.base import BaseCommand
from mainapp.models import ProductCategory, Product
from django.contrib.auth.models import User

import os
import openpyxl

EXEL_PATH = 'mainapp/exel'
FILE_NAME = 'exel_load.xlsx'

# Подгрузка базы из exel
#  функция загркузки продуктов xlsx файла
def load_products_xlsx(file):
    full_path = os.path.join(EXEL_PATH, file)
    wb = openpyxl.load_workbook(full_path)
    sheet = wb['products']
    return sheet

# функция подсчета строк и столбцов в табличке
def cell_coint(sheet):
    i = 1  # определяем число столцов в табице
    j = 1  # определяем число строк
    while True:
        field_name = sheet.cell(row=1, column=i).value
        if not field_name is None:
            i += 1
        else:
            while True:
                row_value = sheet.cell(row=j, column=1).value
                if not row_value is None:
                    j += 1
                else:
                    j = j-1
                    break
            i = i-1
            return i, j

# функция создания списка словарей с данными
def create_db(sheet, i, j):
    products_all = []
    product = []
    new_db = []
    for n1 in range(1, j+1):
        for n2 in range(1, i+1):
            value = sheet.cell(row=n1, column=n2).value
            product.append(value)
        products_all.append(product)
        product = []
    for index, element in enumerate(products_all):
        if index>0:
            dictionary = dict(zip(products_all[0], element))
            new_db.append(dictionary)
    return new_db


class Command(BaseCommand):

    def handle(self, *args, **options):
        xlsx = load_products_xlsx(FILE_NAME)
        i, j = cell_coint(xlsx)
        new_db = create_db(xlsx, i, j)
        Product.objects.all().delete()

        for new in new_db:
            category_name = new['category']
            _c = ProductCategory.objects.get(name=category_name)
            new['category'] = _c
            new_product = Product(**new)
            new_product.save()

    super_user = User.objects.create_superuser('django', 'django@mail.ru', '123456')

